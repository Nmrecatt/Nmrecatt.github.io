import openpyxl
import create_excelhtml 
from datetime import date,datetime

day=int(date.today().day)
mnth=str(date.today().month)
if(int(mnth)<10):
    mnth=str('0'+mnth)
yr=str(date.today().year)
today = str(date.today())

def compute(ws,day,h,c,l,wb,excelpath,xlhtmlpath,time):
    now = datetime.now()
    ws=wb.active
    m_row = ws.max_row 
    ws.cell(row = 1, column = c).value = str(h)+"\n"+time
    ws.column_dimensions['A'].width = 15
    ws.row_dimensions[1].height = 38
    sum=0
    for i in range(2, m_row + 1):
        for j in l:
            if(ws.cell(row = i, column = 1).value == j):
                ws.cell(row = i, column = c).value = 1
        if(ws.cell(row = i, column = c).value == None):
            ws.cell(row = i, column = c).value = 0
        sum+=int(ws.cell(row = i, column = c).value)
    if(day > 0 and day < 16):
        wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        create_excelhtml.view_attendance(cur_sheetpath,xlhtmlpath)
        l=create_excelhtml.view_absentees(cur_sheetpath,c) # c=h+1
    else:
        wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        create_excelhtml.view_attendance(cur_sheetpath,xlhtmlpath)
        l=create_excelhtml.view_absentees(cur_sheetpath,c) # c=h+1
    return sum,l
def create_sheet(l,h,excelpath,xlhtmlpath,time):    
    h=float(h)
    h=int(h)
    #creating a new excel sheet
    if(day == 1 or day == 16):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = today
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        m_row = ws_temp.max_row
        for i in range(1, m_row + 1):
            for j in range(1,m_column+1):
                ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
        wb.save(excelpath+today+".xlsx")
        sum,l=compute(ws,day,h,h+m_column,l,wb,excelpath,xlhtmlpath,time)
     
    #loading an excelsheet if date < 16 
     
    elif(day > 1 and day < 16):
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        sum,l=compute(ws,day,h,h+m_column,l,wb,excelpath,xlhtmlpath,time)
        
    #loading a excel sheet if date  > 16    
        
    else:
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames 
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        sum,l=compute(ws,day,h,h+m_column,l,wb,excelpath,xlhtmlpath,time)
    return sum,l






