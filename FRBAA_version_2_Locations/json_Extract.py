import os
import openpyxl
import json
def json_extractor(dep,year,sub,default,path):
	m=""
	if os.path.exists(default+dep+'/'+year+'/original.json'):
		org_json = open(default+dep+'/'+year+'/original.json','r')
	else:
		m="Face Encodings of all students of "+dep+"-"+year+" year are missing"
		return m 
	data = json.load(org_json)
	if os.path.exists(path+'excel/template.xlsx'):
		wb = openpyxl.load_workbook(path+'excel/template.xlsx')
	else:
		m="Excelsheet template of "+dep+"-"+year+" year "+sub+" subject is missing"
		return m
	ws = wb.active
	m_row = ws.max_row
	dictionary={}
	for i in range(2,m_row+1):
	    for key in data:
	        roll_no = key[0:10]
	        if(ws.cell(row = i, column = 1).value == roll_no):
	            temp_dictionary = {roll_no : data[key]}
	            dictionary.update(temp_dictionary)
	json_path=str(path+"original.json")
	json_file=open(json_path,"w")
	d=json.dumps(dictionary)
	json_file.write(d)
	org_json.close()
	return m 
