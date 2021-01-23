import pandas as pd
import openpyxl

def view_attendance_date(date,excelpath,xlhtmlpath):
	date=str(date)
	yr=date[0:4]
	mnth=date[5:7]
	day=int(date[8:])
	print(yr+","+mnth+","+str(day))
	if(day < 16):
		try:
			wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-01.xlsx")
			ws = wb.active
			sheet = str(wb[date])
			d = sheet[12:22]
			re=pd.read_excel(excelpath+yr+"-"+mnth+"-01.xlsx",sheet_name=date)
			re.to_html(xlhtmlpath)
			return True,d,""
		except:
			return False,0,"Given date may not be a working day"
	else:
		try:
			wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-16.xlsx")
			ws = wb.active
			sheet = str(wb[date])
			d = sheet[12:22]
			re=pd.read_excel(excelpath+yr+"-"+mnth+"-16.xlsx",sheet_name=date)
			re.to_html(xlhtmlpath)
			return True,d,""
		except:
			return False,0,"Given date may not be a working day"

def view_attendance(cur_sheetpath,xlhtmlpath):
	wb=pd.read_excel(cur_sheetpath)
	wb.to_html(xlhtmlpath)

def view_absentees(cur_sheetpath,h):
	l=[]
	l.clear()
	wb=openpyxl.load_workbook(cur_sheetpath)
	ws=wb.active
	m_row=ws.max_row
	for i in range(2,m_row+1):
		if ws.cell(column=h,row=i).value == 0:
			l.append(ws.cell(column=1,row=i).value)
	return l
